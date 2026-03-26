"""File upload — extract text, anonymize, store for chat context."""

from __future__ import annotations

import io
import json
import logging
import uuid

from fastapi import APIRouter, Depends, File, Header, HTTPException, UploadFile

from app.auth import require_workspace
from app.engine.pipeline import PrivacyPipeline
from app.storage import KVStore, get_store
from app import workspace as ws_ops

logger = logging.getLogger("securellm.files")

router = APIRouter(prefix="/v1", tags=["files"])

# Max file size: 10MB
MAX_FILE_SIZE = 20 * 1024 * 1024
# File context TTL: 24 hours
FILE_TTL = 86400


def extract_text_txt(content: bytes, filename: str) -> str:
    """Extract text from plain text files (.txt, .csv, .md, .json, .xml, etc.)."""
    return content.decode("utf-8", errors="replace")


def extract_text_docx(content: bytes) -> str:
    """Extract text from .docx files."""
    from docx import Document
    doc = Document(io.BytesIO(content))
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())


def extract_text_pdf(content: bytes) -> str:
    """Extract text from .pdf files."""
    from pypdf import PdfReader
    reader = PdfReader(io.BytesIO(content))
    texts = []
    for page in reader.pages:
        text = page.extract_text()
        if text:
            texts.append(text)
    return "\n".join(texts)


def extract_text_xlsx(content: bytes) -> str:
    """Extract text from .xlsx files."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    texts = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None]
            if cells:
                texts.append(" | ".join(cells))
    return "\n".join(texts)


def extract_text_pptx(content: bytes) -> str:
    """Extract text from .pptx files."""
    import zipfile
    import re
    zf = zipfile.ZipFile(io.BytesIO(content))
    texts = []
    slide_files = sorted([n for n in zf.namelist() if n.startswith("ppt/slides/slide") and n.endswith(".xml")])
    for slide in slide_files:
        xml = zf.read(slide).decode("utf-8", errors="replace")
        # Extract text from <a:t> tags
        for match in re.finditer(r"<a:t>(.*?)</a:t>", xml):
            t = match.group(1).strip()
            if t:
                texts.append(t)
    return "\n".join(texts)


TEXT_EXTENSIONS = {".txt", ".md", ".csv", ".json", ".xml", ".yaml", ".yml", ".log", ".html", ".htm", ".js", ".py", ".ts", ".sql", ".ini", ".cfg", ".env.example"}


def extract_text(content: bytes, filename: str) -> str:
    """Route to the right extractor based on file extension."""
    name = filename.lower()
    if name.endswith(".docx"):
        return extract_text_docx(content)
    elif name.endswith(".pptx") or name.endswith(".ppt"):
        return extract_text_pptx(content)
    elif name.endswith(".pdf"):
        return extract_text_pdf(content)
    elif name.endswith(".xlsx") or name.endswith(".xls"):
        return extract_text_xlsx(content)
    else:
        # Try as text for known extensions or small files
        ext = "." + name.rsplit(".", 1)[-1] if "." in name else ""
        if ext in TEXT_EXTENSIONS or len(content) < 100_000:
            try:
                return content.decode("utf-8", errors="strict")
            except UnicodeDecodeError:
                pass
    raise ValueError(f"Unsupported file type: {filename}")


@router.post("/upload")
async def upload_file(
    file: UploadFile = File(...),
    x_api_key: str = Header(..., alias="X-API-Key"),
    workspace_id: str = Depends(require_workspace),
    store: KVStore = Depends(get_store),
):
    """Upload a file, extract text, anonymize it, and store for chat context.

    Returns a file_id that can be referenced in chat messages.
    """
    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(413, "File too large (max 20MB)")

    if not file.filename:
        raise HTTPException(400, "Filename required")

    # Extract text
    try:
        raw_text = extract_text(content, file.filename)
    except ValueError as e:
        raise HTTPException(415, str(e))

    if not raw_text.strip():
        raise HTTPException(422, "No text content found in file")

    # Truncate very long documents
    if len(raw_text) > 50_000:
        raw_text = raw_text[:50_000] + "\n\n[Document truncated at 50,000 characters]"

    # Anonymize
    pipeline = await PrivacyPipeline.for_workspace(store, workspace_id)
    anonymized_text, mapping_id = await pipeline.anonymize(raw_text)
    await ws_ops.increment_stats(store, workspace_id)

    # Store file context
    file_id = f"file:{workspace_id}:{uuid.uuid4().hex[:10]}"
    file_data = {
        "filename": file.filename,
        "size": len(content),
        "char_count": len(raw_text),
        "anonymized_text": anonymized_text,
        "mapping_id": mapping_id,
    }
    await store.set(file_id, json.dumps(file_data), ex=FILE_TTL)

    # Store raw file bytes for translation (base64 encoded)
    import base64
    await store.set(f"{file_id}:raw", base64.b64encode(content).decode(), ex=FILE_TTL)

    logger.info("File uploaded: %s (%d bytes, %d chars)", file.filename, len(content), len(raw_text))

    return {
        "file_id": file_id,
        "filename": file.filename,
        "size": len(content),
        "char_count": len(raw_text),
        "preview": anonymized_text[:200] + "..." if len(anonymized_text) > 200 else anonymized_text,
    }


@router.get("/files/{file_id:path}")
async def get_file_info(
    file_id: str,
    workspace_id: str = Depends(require_workspace),
    store: KVStore = Depends(get_store),
):
    """Get file metadata (not the content — content is only used in chat context)."""
    if not file_id.startswith(f"file:{workspace_id}:"):
        raise HTTPException(403, "File does not belong to this workspace")

    raw = await store.get(file_id)
    if not raw:
        raise HTTPException(404, "File not found or expired")

    data = json.loads(raw)
    return {
        "file_id": file_id,
        "filename": data["filename"],
        "size": data["size"],
        "char_count": data["char_count"],
    }
